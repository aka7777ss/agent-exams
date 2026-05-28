#!/usr/bin/env python3
"""Import Agent Arena tasks from CSV/XLSX or prepare PDF image review assets.

CSV/XLSX input columns:
- ID or id
- 题型 or type
- 题目 or prompt
- GT or answer
- 分类/category is optional

PDF input is image-heavy in the current source. The script extracts embedded
page images and writes a review manifest instead of guessing answers.
"""

from __future__ import annotations

import argparse
import csv
import json
import re
from pathlib import Path


TYPE_MAP = {
    "单选": "single_choice",
    "多选": "multiple_choice",
    "json": "json",
    "JSON": "json",
    "数字": "number",
    "数值": "number",
    "number": "number",
    "short_text": "short_text",
    "文本": "short_text",
}


def normalize_key(value: str) -> str:
    return str(value or "").strip().lower().replace(" ", "").replace("_", "")


def pick(row: dict, names: list[str]) -> str:
    normalized = {normalize_key(key): value for key, value in row.items()}
    for name in names:
        value = normalized.get(normalize_key(name))
        if value is not None and str(value).strip():
            return str(value).strip()
    return ""


def parse_options(prompt: str) -> list[dict]:
    matches = list(re.finditer(r"(?<![A-Za-z])([A-D])(?:[.、．])\s*", prompt))
    if len(matches) < 2:
        matches = list(re.finditer(r"(?<![A-Za-z])([A-D])(?:\s+)", prompt))
    if len(matches) < 2:
        return []

    options = []
    for index, match in enumerate(matches):
        start = match.end()
        end = matches[index + 1].start() if index + 1 < len(matches) else len(prompt)
        text = prompt[start:end].strip().strip(" ：:；;，,").strip()
        options.append({"id": match.group(1), "text": text})
    return options


def prompt_without_options(prompt: str) -> str:
    first_option = re.search(r"(?<![A-Za-z])A(?:[.、．])\s*", prompt) or re.search(r"(?<![A-Za-z])A(?:\s+)", prompt)
    if not first_option:
        return prompt.strip()
    return prompt[: first_option.start()].strip().strip(" ：:；;，,").strip()


def parse_gt(raw: str, task_type: str):
    value = raw.strip()

    if task_type == "multiple_choice":
        if value.startswith("["):
            return json.loads(value)
        return [item.strip() for item in re.split(r"[,，、]", value) if item.strip()]

    if task_type == "json":
        return json.loads(value)

    if task_type == "number":
        try:
            number = float(value)
            return int(number) if number.is_integer() else number
        except ValueError:
            return value

    return value


def build_schema(task_type: str, options: list[dict]) -> dict:
    if task_type == "single_choice":
        return {"type": "string", "enum": [option["id"] for option in options] or ["A", "B", "C", "D"]}
    if task_type == "multiple_choice":
        return {"type": "array", "items": {"type": "string"}}
    if task_type == "json":
        return {"type": "object"}
    if task_type == "number":
        return {"type": "number"}
    return {"type": "string"}


def build_task(row: dict, index: int) -> dict:
    raw_id = pick(row, ["ID", "id", "题号"]) or f"q{index:03d}"
    raw_type = pick(row, ["题型", "type"])
    prompt = pick(row, ["题目", "prompt", "question"])
    gt = pick(row, ["GT", "answer", "答案"])
    module = pick(row, ["module", "模块"]) or "general"
    category = pick(row, ["category", "分类", "赛道"]) or raw_id[:1] or "general"
    category_path = pick(row, ["category_path", "分类路径"]) or f"{module}/{category}"
    capability_tags = [item.strip() for item in (pick(row, ["capability_tags", "capability", "能力标签"]) or "").replace("，", ",").split(",") if item.strip()]

    task_type = TYPE_MAP.get(raw_type, TYPE_MAP.get(raw_type.strip(), raw_type.strip()))
    if task_type not in {"single_choice", "multiple_choice", "json", "number", "short_text"}:
        raise ValueError(f"Unsupported type for {raw_id}: {raw_type}")
    if not prompt:
        raise ValueError(f"Missing prompt for {raw_id}")
    if not gt:
        raise ValueError(f"Missing GT for {raw_id}")

    options = parse_options(prompt) if task_type in {"single_choice", "multiple_choice"} else []
    public_prompt = prompt_without_options(prompt) if options else prompt
    answer = parse_gt(gt, task_type)

    task = {
        "id": raw_id,
        "module": module,
        "category": category,
        "category_path": category_path,
        "type": task_type,
        "prompt": public_prompt,
        "answer_schema": build_schema(task_type, options),
        "grader": {"type": "exact_match"},
    }
    if capability_tags:
        task["capability_tags"] = capability_tags

    if options:
        task["options"] = options
    if task_type == "number" and isinstance(answer, (int, float)):
        task["grader"]["type"] = "number_range"
        task["grader"].update({"min": answer, "max": answer})
    elif task_type == "json" and isinstance(answer, dict):
        task["grader"]["type"] = "json_fields"
        task["grader"]["required_fields"] = answer
    else:
        task["grader"]["answer"] = answer

    return task


def read_csv(path: Path) -> list[dict]:
    dialect = "excel-tab" if path.suffix.lower() == ".tsv" else "excel"
    with path.open("r", encoding="utf-8-sig", newline="") as file:
        return list(csv.DictReader(file, dialect=dialect))


def read_xlsx(path: Path) -> list[dict]:
    try:
        from openpyxl import load_workbook
    except ModuleNotFoundError as error:
        raise SystemExit("XLSX import requires openpyxl. Install it or use a CSV export.") from error

    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    headers = [str(cell or "").strip() for cell in rows[0]]
    return [
        {headers[index]: "" if value is None else value for index, value in enumerate(row)}
        for row in rows[1:]
        if any(value is not None and str(value).strip() for value in row)
    ]


def validate_tasks(tasks: list[dict]) -> None:
    ids = set()
    for task in tasks:
        if task["id"] in ids:
            raise ValueError(f"Duplicate task id: {task['id']}")
        ids.add(task["id"])
        if "grader" not in task:
            raise ValueError(f"Missing grader: {task['id']}")
        if task["type"] in {"single_choice", "multiple_choice"} and not task.get("options"):
            raise ValueError(f"Missing options: {task['id']}")


def import_table(input_path: Path, output_path: Path) -> None:
    rows = read_xlsx(input_path) if input_path.suffix.lower() in {".xlsx", ".xlsm"} else read_csv(input_path)
    tasks = [build_task(row, index + 1) for index, row in enumerate(rows)]
    validate_tasks(tasks)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(json.dumps({"tasks": tasks}, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(f"Wrote {len(tasks)} tasks to {output_path}")


def extract_pdf_images(input_path: Path, output_dir: Path) -> None:
    try:
        from pypdf import PdfReader
    except ModuleNotFoundError as error:
        raise SystemExit("PDF review import requires pypdf. Install it or run with the bundled Codex Python runtime.") from error

    output_dir.mkdir(parents=True, exist_ok=True)
    reader = PdfReader(str(input_path))
    pages = []

    for page_number, page in enumerate(reader.pages, 1):
        page_images = []
        for image_index, image in enumerate(page.images, 1):
            image_path = output_dir / f"page_{page_number:02d}_image_{image_index:02d}_{image.name}"
            image_path.write_bytes(image.data)
            page_images.append(str(image_path))
        pages.append({"page": page_number, "images": page_images, "text_layer": page.extract_text() or ""})

    manifest = {
        "source": str(input_path),
        "note": "PDF contains image-heavy tables. Review images/OCR output before replacing data/tasks.json.",
        "pages": pages,
    }
    manifest_path = output_dir / "review_manifest.json"
    manifest_path.write_text(json.dumps(manifest, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(f"Extracted review assets to {output_dir}")
    print(f"Review manifest: {manifest_path}")


def main() -> int:
    parser = argparse.ArgumentParser(description="Import Agent Arena tasks.")
    parser.add_argument("input", type=Path, help="CSV/XLSX/PDF path")
    parser.add_argument("--output", type=Path, default=Path("data/tasks.json"))
    parser.add_argument("--pdf-review-dir", type=Path, default=Path("data/import_review"))
    args = parser.parse_args()

    suffix = args.input.suffix.lower()
    if suffix == ".pdf":
        extract_pdf_images(args.input, args.pdf_review_dir)
    elif suffix in {".csv", ".tsv", ".xlsx", ".xlsm"}:
        import_table(args.input, args.output)
    else:
        raise SystemExit(f"Unsupported input type: {suffix}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
